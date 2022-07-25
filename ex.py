# Python 3 server example
from http.server import BaseHTTPRequestHandler, HTTPServer
from openpyxl import load_workbook
import cgi
import datetime
import calendar
from datetime import timedelta
import xlsxwriter
import pandas as pd
import numpy as np
hostName = "localhost"
serverPort = 5557
showThem=[]
tts=[]
dates=[]
cur_tt=[0]
selected=[]
seldetails=[]
curTeachers=[]
seldates=[]
tempdisp=[]
df = pd.DataFrame(pd.read_excel("final.xlsx"))
df=df.astype(str,errors='raise')

c=['Date','Day','Subject', 'P1','P2', 'P3', 'SB', 'P4', 'P5', 'P6', 'LB','P7', 'P8', 'P9']
workbook = load_workbook(filename="tts.xlsx")
sheets=workbook.sheetnames
for i in range(len(sheets)):
    t=[]
    tt=[]
    tt.append([sheets[i],'nan','nan','nan','nan','nan','nan','nan','nan','nan','nan','nan','nan','nan'])
    dict=pd.read_excel('tts.xlsx',sheet_name=sheets[i])
    for j in dict:
        t.append(dict[j])
    if(len(t)>0):
        for k in range(len(t[0])):
            trow=[]
            for l in range(1,15):
                trow.append(t[l][k])
            #seldates.append(t[1][k]+"-"+t[2][k])
            tt.append(trow)
        ftt=pd.DataFrame(tt)
        ftt.columns = c
        tts.append(ftt)


class MyServer(BaseHTTPRequestHandler):  
    # GET
    def do_GET(self):
        if self.path.endswith('/start'):

            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            h=open("selteacher.html","rb")
            self.wfile.write(h.read())
            for i in range(len(seldates)):
                self.wfile.write(bytes("<option value=\""+seldates[i]+"\">"+seldates[i]+"</option>", "utf-8"))
            self.wfile.write(bytes("</select></div></div>", "utf-8"))
            if len(showThem)!=0:
                self.wfile.write(bytes(showThem[0], "utf-8"))
            self.wfile.write(bytes("</body></html>", "utf-8"))
        elif self.path.endswith('/seldate'):
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            h=open("seldate.html","rb")
            self.wfile.write(h.read())
            self.wfile.write(bytes("</body></html>", "utf-8"))
        elif self.path.endswith('/home'):
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            h=open("home.html","rb")
            self.wfile.write(h.read())
            self.wfile.write(bytes("</body></html>", "utf-8"))
        elif self.path.endswith('/showtts'):
            display=""
            for i in range(len(sheets)):
                display+="<div class=\"i"+str(i)+"\"><button class = \"selbut\" id=\""+str(i)+"\" name = \"id"+str(i)+"\" type=\"submit\" value=\"Submit\">"+sheets[i]+"</button></div>"
            display+="</div></form></div>"
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            h=open("showtts.html","rb")
            self.wfile.write(h.read())
            self.wfile.write(bytes(display, "utf-8"))
            self.wfile.write(bytes("</body></html>", "utf-8"))
        elif self.path.endswith('/showtt'):
            a=0
            c=['Date','Day','Subject', 'P1','P2', 'P3', 'SB', 'P4', 'P5', 'P6', 'LB','P7', 'P8', 'P9']
            d=tts[cur_tt[0]]
            display=""
            display+="<div class=\"abcd\">"
            display+="<div class =\"mtext\"><h>"+d.iloc[0][0]+"</h></div>"
            display+="<table><tr>"
            for i in range(len(c)):
                display+="<td>"+c[i]+"</td>"
            display+="</tr>"
            for i in d.iterrows():
                if a!=0:
                    display+="<tr><td>"+i[1].Date+"</td><td>"+i[1].Day+"</td><td>"+i[1].Subject+"</td><td>"+i[1].P1+"</td><td>"+i[1].P2+"</td><td>"+i[1].P3+"</td><td>"+i[1].SB+"</td><td>"+i[1].P4+"</td><td>"+i[1].P5+"</td><td>"+i[1].P6+"</td><td>"+i[1].LB+"</td><td>"+i[1].P7+"</td><td>"+i[1].P8+"</td><td>"+i[1].P9+"</td></tr>"
                a+=1
            display+="<form id=\"frm1\" method=\"POST\" enctype=\"multipart/form-data\" ><div class=\"backButton\"><button class = \"but\" id=\"get_stock\" type=\"submit\" value=\"Submit\" name=\"backButton\">BACK</button></div></form></div>"
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            h=open("showtt.html","rb")
            self.wfile.write(h.read())
            self.wfile.write(bytes(display, "utf-8"))
            self.wfile.write(bytes("</body></html>", "utf-8"))


    # POST

    def do_POST(self):
        if self.path.endswith('/start'):
            allowed=0
            print("Inside Start")
            ctype, pdict = cgi.parse_header(self.headers.get('content-type')) 
            pdict['boundary'] = bytes (pdict['boundary'], "utf-8")
            content_len = int(self.headers.get('Content-length'))
            pdict['CONTENT-LENGTH'] = content_len 
            if (True):
                print("INside IF!!!")
                z=10
                fields = cgi.parse_multipart(self.rfile, pdict)
                try:
                    f=fields.get('showTeachers')[0]
                    z=1
                except:
                    z=10
                try:
                    t=fields.get('showtt')[0]
                    z=0
                except:
                    z=z
                
                
                if z==1:
                    if ctype == 'multipart/form-data':
                        c=0
                        print("Inside If !!!")
                        allowed=1
                        try:
                            p=fields.get('sel_period')[0]
                            d=fields.get('sel_day')[0]
                        except:
                            allowed=0
                        print(p,d)
                        if allowed==1 and p!="Select Period" and d!="Select Day":
                            x=df.loc[df.Day==d[11:]]
                            a=int(p[-1])
                            if a==1:
                                y=x.loc[x.P1=='1']
                            elif a==2:
                                y=x.loc[x.P2=='1']
                            elif a==3:
                                y=x.loc[x.P3=='1']
                            elif a==4:
                                y=x.loc[x.P4=='1']
                            elif a==5:
                                y=x.loc[x.P5=='1']
                            elif a==6:
                                y=x.loc[x.P6=='1']
                            elif a==7:
                                y=x.loc[x.P7=='1']
                            elif a==8:
                                y=x.loc[x.P8=='1']
                            elif a==9:
                                y=x.loc[x.P9=='1']
                            
                            if len(seldetails)!=0:
                                seldetails.pop()
                                seldetails.pop()
                            seldetails.append(p)
                            seldetails.append(d)
                            while(len(curTeachers)>0):
                                curTeachers.pop()
                            curTeachers.append(y['Name'].tolist())
                            res=y[['Name','Subject']]
                            display=""
                            #display+="<form id=\"frm1\" method=\"GET\" enctype=\"multipart/form-data\">"
                            display+="<div class =\"mtext\"><h>Period : "+p+"  |   Day : "+d+"</h></div>"
                            display+="<table class=\"avail\"><tr><td>Name</td><td>Subject</td><td>Select</td></tr>"
                            for i in res.iterrows():
                                selected.append(0)
                                display+="<tr><th>"+i[1].Name+"</th><th>"+i[1].Subject+"</th><th>"+"<button class = \"selbut\" id=\"get_stock\" name = \"id"+str(c)+"\" type=\"submit\" value=\"Submit\">SELECT</button><button class=\"next\" type=\"submit\" name=\"arr"+str(c)+"\" id=\"arr"+str(c)+"\" >&#10095;</button></th></tr>"
                                c+=1
                            display+="</form></div>"
                            while(len(showThem)>0):
                                showThem.pop()

                            while(len(tempdisp)>0):
                                tempdisp.pop()
                            showThem.append(display)
                            tempdisp.append(display)
                            print(selected)
                        self.send_response(301)
                        self.send_header('Content-type', 'text/html') 
                        self.send_header('Location', '/start')
                        self.end_headers()
                elif z==0:
                    print("Inside show TimeTable!!!")
                    self.send_response(301)
                    self.send_header('Content-type', 'text/html') 
                    self.send_header('Location', '/showtt')
                    self.end_headers()
                elif z==10:
                    allowed2=0
                    print("Inside z=10")
                    x=0
                    for i in range(len(selected)):
                        try:
                            a=fields.get('id'+str(i))[0]
                            x=i
                            allowed=1
                        except:
                            x=x
                        try:
                            a=fields.get('arr'+str(i))[0]
                            x=i
                            allowed2=1
                        except:
                            x=x
                        
                    
                    if allowed==1:
                        selected[x]=1
                        ctt=tts[cur_tt[0]]
                        ctt.loc[ctt["Date"] == seldetails[1][:10], seldetails[0]] = curTeachers[0][x]
                        tts[cur_tt[0]]=ctt
                        # with pd.ExcelWriter('tts.xlsx', engine='xlsxwriter') as writer:
                        #     ctt.iloc[1:].to_excel(writer, sheet_name=ctt.iloc[0][0])
                        with pd.ExcelWriter('tts.xlsx', engine='xlsxwriter') as writer:
                            for i in range(len(tts)):
                                tts[i].iloc[1:].to_excel(writer, sheet_name=tts[i].iloc[0][0])
                        print(curTeachers[0][x],seldetails[1][:10])
                        
                        showThem.pop()
    
                    elif allowed2==1:
                        d = pd.DataFrame(pd.read_excel("ttwclass.xlsx"))
                        c=['Day','P1','P2', 'P3', 'SB', 'P4', 'P5', 'P6', 'LB','P7', 'P8', 'P9']
                        showThem[0]=tempdisp[0]
                        display=showThem[0]
                        print("INside Allowed 2!!!")
                        dtt=d[d.Name==curTeachers[0][x]]
                        
                        display+="<div class =\"tttext\"><h>"+curTeachers[0][x]+" :</h></div><table class=\"tt\"><tr>"
                        for i in range(len(c)):
                            display+="<td>"+c[i]+"</td>"
                        display+="</tr>"
                        for i in dtt.iterrows():
                            display+="<tr><td>"+i[1].Day+"</td><td>"+i[1].P1+"</td><td>"+i[1].P2+"</td><td>"+i[1].P3+"</td><td>"+i[1].SB+"</td><td>"+i[1].P4+"</td><td>"+i[1].P5+"</td><td>"+i[1].P6+"</td><td>"+i[1].LB+"</td><td>"+i[1].P7+"</td><td>"+i[1].P8+"</td><td>"+i[1].P9+"</td></tr>"
                        display+="</table>"
                        showThem[0]=display
                    self.send_response(301)
                    self.send_header('Content-type', 'text/html') 
                    self.send_header('Location', '/start')
                    self.end_headers()



        elif self.path.endswith('/seldate'):
            print('entered if!!!')
            ctype, pdict = cgi.parse_header(self.headers.get('content-type')) 
            pdict['boundary'] = bytes (pdict['boundary'], "utf-8")
            content_len = int(self.headers.get('Content-length'))
            pdict['CONTENT-LENGTH'] = content_len 
            if ctype == 'multipart/form-data':
                allowed=1
                while(len(seldates)>0):
                    seldates.pop()
                fields = cgi.parse_multipart(self.rfile, pdict)
                try:
                    f=fields.get('sel_period')[0]
                    t=fields.get('sel_day')[0]
                    print(f,t)
                except:
                    allowed=0
                if allowed==1:
                    print("Entered Allowed")
                    tname=fields.get('fname')[0]
                    c=['Date','Day','Subject', 'P1','P2', 'P3', 'SB', 'P4', 'P5', 'P6', 'LB','P7', 'P8', 'P9']
                    tt=[]
                    val1=t
                    val2=f
                    dt1 = datetime.datetime(int(val1[:4]),int(val1[5:7]),int(val1[8:]))
                    dt2 = datetime.datetime(int(val2[:4]),int(val2[5:7]),int(val2[8:]))
                    calendar.day_name[dt1.weekday()][0:3]
                    diff=dt2-dt1
                    temp=dt1
                    tt.append([tname,'nan','nan','nan','nan','nan','nan','nan','nan','nan','nan','nan','nan','nan'])
                    if(diff.days>0):
                        for i in range(diff.days+1):
                            ttentry=[]
                            sd=""
                            sdate=temp.strftime("%d/%m/%Y")
                            dayname=calendar.day_name[temp.weekday()][0:3]
                            sd+=sdate+"-"+dayname
                            ttentry.append(sdate)
                            ttentry.append(dayname)
                            ttentry.append('-')
                            for i in range(11):
                                ttentry.append('-')
                            tt.append(ttentry)
                            seldates.append(sd)
                            temp=temp+ timedelta(days=1)
                        finaltt=pd.DataFrame(tt)
                        finaltt.columns = c
                        
                        tts.append(finaltt)
                        with pd.ExcelWriter('tts.xlsx', engine='xlsxwriter') as writer:
                                    for i in range(len(tts)):
                                        tts[i].iloc[1:].to_excel(writer, sheet_name=tts[i].iloc[0][0])
                        cur_tt[0]=len(tts)-1

                        self.send_response(301)
                        self.send_header('Content-type', 'text/html') 
                        self.send_header('Location', '/start')
                        self.end_headers()
                    else :
                        self.send_response(301)
                        self.send_header('Content-type', 'text/html') 
                        self.send_header('Location', '/seldate')
                        self.end_headers()
                else :
                    self.send_response(301)
                    self.send_header('Content-type', 'text/html') 
                    self.send_header('Location', '/seldate')
                    self.end_headers()


        elif self.path.endswith('/home'):
            print('entered if!!!')
            ctype, pdict = cgi.parse_header(self.headers.get('content-type')) 
            pdict['boundary'] = bytes (pdict['boundary'], "utf-8")
            content_len = int(self.headers.get('Content-length'))
            pdict['CONTENT-LENGTH'] = content_len 
            if ctype == 'multipart/form-data':
                x=0
                fields = cgi.parse_multipart(self.rfile, pdict)
                try:
                    f=fields.get('new')[0]
                except:
                    x=0
                try:
                    t=fields.get('old')[0]
                except:
                    x=1
                if x==1:
                    self.send_response(301)
                    self.send_header('Content-type', 'text/html') 
                    self.send_header('Location', '/seldate')
                    self.end_headers()
                else:
                    self.send_response(301)
                    self.send_header('Content-type', 'text/html') 
                    self.send_header('Location', '/showtts')
                    self.end_headers()
        elif self.path.endswith('/showtts'):
            print("Entered showtts")
            ctype, pdict = cgi.parse_header(self.headers.get('content-type')) 
            pdict['boundary'] = bytes (pdict['boundary'], "utf-8")
            content_len = int(self.headers.get('Content-length'))
            pdict['CONTENT-LENGTH'] = content_len 
            if ctype == 'multipart/form-data':
                fields = cgi.parse_multipart(self.rfile, pdict)
                print("Entered IF!!!")
                x=10
                allowed=0
                for i in range(len(sheets)):
                    print("Entered Loop!!")
                    print(sheets)
                    try:
                        a=fields.get('id'+str(i))[0]
                        x=i
                        allowed=1
                        print("allowed!!!")
                    except:
                        x=x
                if allowed==1:
                    cur_tt[0]=x
                    while (len(seldates)>0):
                        seldates.pop()
                    for i in range(1,len(tts[cur_tt[0]].index)):
                        seldates.append(tts[cur_tt[0]]['Date'][i]+"-"+tts[cur_tt[0]]['Day'][i])
                    print(seldates)
                
            self.send_response(301)
            self.send_header('Content-type', 'text/html') 
            self.send_header('Location', '/start')
            self.end_headers()
        elif self.path.endswith('/showtt'):
            self.send_response(301)
            self.send_header('Content-type', 'text/html') 
            self.send_header('Location', '/start')
            self.end_headers()

if __name__ == "__main__":        
    webServer = HTTPServer((hostName, serverPort), MyServer)
    print("Server started http://%s:%s" % (hostName, serverPort))

    try:
        webServer.serve_forever()
    except KeyboardInterrupt:
        pass

    webServer.server_close()
    print("Server stopped.")